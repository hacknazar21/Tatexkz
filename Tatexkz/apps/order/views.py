from cmath import pi
from datetime import datetime
import json
import locale
from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
import openpyxl
from django.views.decorators.csrf import csrf_exempt
import xml.etree.ElementTree as ET
import requests
from Tatexkz.apps.order.models import Tariff

from Tatexkz.apps.promo.models import Promo

with open("static/files/countries.min.json", "r", encoding='utf-8') as read_file:
    countries = json.load(read_file)

countiesRoot = ET.parse('static/files/counties-codes.xml')
countriesCodes = {}
for country in countiesRoot.iter('country'):
    for name in country.iter('name'):
        for alpha2 in country.iter('alpha2'):
            countriesCodes[name.text] = alpha2.text
@csrf_exempt
def tariff(request):
    if request.method == "POST":
        jsonReq = json.loads(request.body.decode())
        typePackage = jsonReq.get('type', '')
        fromCity = jsonReq.get('from', '')
        whereCity = jsonReq.get('where', '')
        weight = jsonReq.get('weight', '')
        fromCountry = jsonReq.get('from_country', '')
        whereCountry = jsonReq.get('where_country', '')
        promo = jsonReq.get('promo', '')
        postIndexSender = jsonReq.get('postIndexSender', '000000')
        postIndexRecipient = jsonReq.get('postIndexRecipient', '000000')
        DlvyDateTime = ''
        promoResp = calcPromo(promo)
        """ if(fromCountry != ''):
            fromCountryCode = countriesCodes[fromCountry]
            whereCountryCode = countriesCodes[whereCountry]
            dateSend = datetime.strptime(jsonReq.get(
                'date', ''), '%d/%m/%Y').strftime("%Y-%m-%d")
            
            tree = ET.parse('static/files/req_date.xml')
            root = tree.getroot()
            i = 0
            for ShipmentWeight in root.iter('ShipmentWeight'):
                ShipmentWeight.text = str(weight)
            for Date in root.iter('Date'):
                Date.text = dateSend
            for CountryCode in root.iter('CountryCode'):
                if not i:
                    CountryCode.text = fromCountryCode
                else:
                    CountryCode.text = whereCountryCode
                    i = 0
                    break
                i += 1
            postIndexSender = ''.join(i for i in postIndexSender if not i.isalpha())
            postIndexRecipient = ''.join(i for i in postIndexRecipient if not i.isalpha())
            for PostalCode in root.iter('Postalcode'):
                if not i:
                    PostalCode.text = str(postIndexSender)
                else:
                    PostalCode.text = str(postIndexRecipient)
                    i = 0
                    break
                i += 1
            tree.write('static/files/req_date.xml', 'UTF-8')
            with open('static/files/req_date.xml') as inputfile:
                xml_file = inputfile.read()
            response = requests.post(
                'http://xmlpi-ea.dhl.com/XMLShippingServlet?isUTF8Support=true', data=xml_file)
            tree = ET.ElementTree(ET.fromstring(response.text))

            root = tree.getroot()
            for DlvyDateTime in root.iter('DlvyDateTime'):
                DlvyDateTime = DlvyDateTime.text
            
            DlvyDateTime = datetime.strptime(DlvyDateTime, 
                        '%Y-%m-%d %H:%M:%S').strftime("%d.%m.%Y") """
        if fromCountry == '' and whereCountry == '':
            for country in countries:
                for city in countries[country]:
                    if city == fromCity:
                        fromCountry = country
                    if city == whereCity:
                        whereCountry = country
                    if whereCountry != '' and fromCountry != '':
                        break
                if whereCountry != '' and fromCountry != '':
                    break
            if whereCountry != 'Казахстан' and fromCountry != 'Казахстан':
                return JsonResponse({'error':  True, 'errorText': "Упс, можно только из Казахстана или в него"})
            if whereCountry == '':
                return JsonResponse({'error':  True, 'errorText': 'Не найдена страна с городом ' + whereCity})
            elif fromCountry == '':
                return JsonResponse({'error':  True, 'errorText': 'Не найдена страна с городом ' + fromCity})
        calculatedTariff = calcTariff(
            typePackage, fromCity, whereCity, fromCountry, whereCountry, weight)
        oldPrice = calculatedTariff

        if promoResp['percent'] > -1:
            if isint(calculatedTariff):
                calculatedTariff = calculatedTariff - \
                    calculatedTariff*(promoResp['percent']/100)
        else:
            return JsonResponse({'error':  True, 'errorText': promoResp['error']})
        if isint(calculatedTariff):
            return JsonResponse({'error':  False, 'oldPrice': oldPrice, 'price':  calculatedTariff, 'percent': promoResp['percent'], 'DlvyDateTime': 'DlvyDateTime'})
        else:
            return JsonResponse({'error':  True, 'errorText': calculatedTariff})


def isint(s):
    try:
        int(s)
        return True
    except ValueError:
        return False


def calcTariff(typePackage, fromcity, wherecity, fromCountry, whereCountry, weight):
    if fromcity == wherecity:
        if weight <= 3:
            return 1650
        elif weight <= 7:
            return 2950
        elif weight <= 10:
            return 4500
        else:
            return int(4500 + ((weight - 10)/0.5)*370)
    if fromCountry == 'Казахстан':
        tariffType = 'export'
    else:
        tariffType = 'import'
    if weight > 300:
        return 'Слишком большой груз'
    tariffFile = Tariff.objects.latest('tariffFile')
    wb_obj = openpyxl.load_workbook(tariffFile.tariffFile.path)
    zonesSheet = wb_obj.get_sheet_by_name("zones")
    zones = {}
    zone = ''
    for i, row in enumerate(zonesSheet.iter_rows(values_only=True)):
        if i != 0 and row[0] != None and row[1] != None:
            zones[row[0]] = []
            zones[row[0]].append(row[1])
    if(fromCountry == whereCountry):
        if weight > 50:
            return 'Слишком большой груз'
        ieSheet = wb_obj.get_sheet_by_name('kz')
        weightMass = []
        for i in range(1, ieSheet.max_row):
            weightMass.append(ieSheet.cell(row=i, column=1).value)
        weight = normalizeWeight(weightMass, weight)
        resultRow = weightMass.index(weight) + 1

        return ieSheet.cell(row=resultRow, column=2).value
    elif(tariffType == 'import'):
        zone = zones.get(
            fromCountry, ['В данную страну нет доставки: ' + fromCountry])[0]
    else:
        zone = zones.get(whereCountry, [
            'В данную страну нет доставки: ' + whereCountry])[0]
    if not isint(zone):
        return zone
    ieSheet = wb_obj.get_sheet_by_name(tariffType)
    resultColumn = 1
    resultRow = 1

    for i in range(2, 9):
        if ieSheet.cell(row=2, column=i).value == zone:
            print(i)
            resultColumn = i
    print(typePackage)
    if typePackage == 'document' and weight < 2:
        
        weightMass = []
        for i in range(3, 6):
            weightMass.append(ieSheet.cell(row=i, column=1).value)
        weight = normalizeWeight(weightMass, weight)
        resultRow = weightMass.index(weight) + 3
    else:
        weightMass = []
        for i in range(8, ieSheet.max_row):
            weightMass.append(ieSheet.cell(row=i, column=1).value)
        weight = normalizeWeight(weightMass, weight)

        resultRow = weightMass.index(weight) + 8

    return ieSheet.cell(row=resultRow, column=resultColumn).value


def normalizeWeight(weightMass, weight):
    for currentWeight in weightMass:
        if(weight <= currentWeight):
            return currentWeight


@csrf_exempt
def index(request):
    type = ''
    weight = ''
    fromcity = ''
    wherecity = ''
    fromcountry = ''
    wherecountry = ''
    height = ''
    width = ''
    length = ''
    promo = ''
    if(request.GET):
        try:
            fromcity = request.GET.get('from', '')
            wherecity = request.GET.get('where', '')
            type = request.GET.get('type', '')
            promo = request.GET.get('promo', '')

            if(request.GET.get('weight', 'Не найдено').find('_') != -1):
                weight = float(request.GET.get(
                    'weight', 'Не найдено').replace('._ (кг)', ''))
            else:
                weight = float(request.GET.get(
                    'weight', 'Не найдено').replace(' (кг)', ''))
            if request.GET.get('height', '') != '':
                height = int(request.GET.get(
                    'height', 'Не найдено').replace(' (см)', ''))
                width = int(request.GET.get(
                    'width', 'Не найдено').replace(' (см)', ''))
                length = int(request.GET.get(
                    'length', 'Не найдено').replace(' (см)', ''))
        except ValueError:
            return HttpResponse("Ой. Что-то пошло не так")

    if type != '' and weight != '' and fromcity != '' and wherecity != '':
        for country in countries:
            for city in countries[country]:
                if city == fromcity:
                    fromcountry = country
                if city == wherecity:
                    wherecountry = country
                if wherecountry != '' and fromcountry != '':
                    break
            if wherecountry != '' and fromcountry != '':
                break
        if wherecountry != 'Казахстан' and fromcountry != 'Казахстан':
            return HttpResponse("Упс, можно только из Казахстана или в него")
        if wherecountry == '':
            return HttpResponse('Не найдена страна с городом ' + wherecity)
        elif fromcountry == '':
            return HttpResponse('Не найдена страна с городом ' + fromcity)

        calculatedTariff = calcTariff(
            type, fromcity, wherecity, fromcountry, wherecountry, weight)
        if isint(calculatedTariff):
            return render(request, 'order/order.html', {'countries': countries,
                                                        'citiesfrom': countries[fromcountry],
                                                        'citieswhere': countries[wherecountry],
                                                        'activeCountryFrom': fromcountry,
                                                        'activeCountryWhere': wherecountry,
                                                        'fromcity': fromcity,
                                                        'wherecity': wherecity,
                                                        'weight': str(weight),
                                                        'height': height,
                                                        'width': width,
                                                        'length': length,
                                                        'type': type,
                                                        'error':  False,
                                                        'price':  calculatedTariff,
                                                        'promo': promo})
        else:
            return render(
                request, 'order/order.html', {'countries': countries,
                                              'citiesfrom': countries[fromcountry],
                                              'citieswhere': countries[wherecountry],
                                              'activeCountryFrom': fromcountry,
                                              'activeCountryWhere': wherecountry,
                                              'fromcity': fromcity,
                                              'wherecity': wherecity,
                                              'weight': str(weight),
                                              'type': type,
                                              'error':  True,
                                              'errorText': calculatedTariff,
                                              'promo': promo}
            )

    return render(
        request, 'order/order.html', {'countries': countries,
                                      'weight': '', 'type': 'document'}
    )


def country(request, country_name):
    if country_name in countries:
        return JsonResponse({'cities': countries[country_name]})
    else:
        return JsonResponse({'cities': 'Error'})


def getcountries(request):
    return JsonResponse(countries)


def calcPromo(promo):
    if not promo == '':
        try:
            promocode = Promo.objects.get(promo=promo)
            if promocode.datefrom.isoformat() > datetime.now().isoformat():
                print("Рано")
                return {'error': 'Промокод еще не активен', 'percent': -1}
            elif datetime.now().isoformat() < promocode.dateto.isoformat():
                print("В самый раз")
                return {'error': '', 'percent': promocode.percent}
            else:
                print("Поздно")
                return {'error': 'Срок действия промокода истёк', 'percent': -1}

        except Promo.DoesNotExist:
            return {'error': 'Данного промокода не существует', 'percent': -1}
    else:
        return {'error': '', 'percent': 0}
